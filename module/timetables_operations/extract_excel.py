from openpyxl import load_workbook
import openpyxl
from module.timetables_operations.times_op import isTimeFormat,isTimeFormatH
import os
import asyncio

path="~/FCE_bot/module/timetables_operations/"

def find_files(tipo:str) -> list:
    arr=[]
    if (tipo=="bus"):
        arr=os.listdir(path+"bus/")
    elif (tipo=="littorina"):
        arr=os.listdir(path+"littorina/")
    try:
        arr.remove("locations.txt")
    except:
        arr
    return arr


bus_workbooks=[]
train_workbooks=[]

def load(tipo:str) -> None:
    for fname in find_files(tipo):
        if tipo=="bus":
            bus_workbooks.append(load_workbook(filename=path+tipo+'/'+fname))
        elif tipo=="littorina":
            train_workbooks.append(load_workbook(filename=path+tipo+'/'+fname))

async def extract(tipo:str,work_index:int) -> list:
    workbook=openpyxl.Workbook
    if tipo=="bus":
        workbook=bus_workbooks[work_index]
        
    elif tipo=="littorina":
        workbook=train_workbooks[work_index]

    matrix=[]
    n=0
    for sheet_name in workbook.sheetnames:
        sheet=workbook[sheet_name]
        n_righe=0
        matrix.append([])
        for row in sheet.iter_rows():
            n_righe+=1

        matrix[n]=[[] for i in range(n_righe)]

        i=0
        for row in sheet.iter_rows(values_only=True):
            for j in row:
                matrix[n][i].append(j)    
            i+=1
        n+=1
    
    return matrix #array di matrici

def dimensions(matr) -> list: #ritorna array con indice di partenza e righe
    rows=len(matr)
    start=0
    for i in range(rows):
        if str(matr[i][0]) in ["CATANIA (S.Sofia)",]: #la stringa deve essere un identificativo per l'indice di partenza
            start=i
            break
    dim=[start,rows]
    return dim

def all_replacing(s:str):
    if s!=None:
        s=s.upper().replace(' ','').replace("P.ZZA","PIAZZA").replace("Ù","U'").replace("'",'').replace('°','').replace('.','').replace('(','').replace(')','').replace('METRO','')
    # if s=="NESIMA":
    #     s="CATANIANESIMA"
    # elif s=="CIBALI":
    #     s="CATANIACIBALI"
    return s


def locations_to_file(tipo:str) -> None:
    different_loc=[]
    toadd=[]
    if tipo=="bus": workbooks=bus_workbooks
    elif tipo=="littorina": workbooks=train_workbooks
    
    for work_index in range(len(workbooks)):
        m_table=asyncio.run(extract(tipo,work_index))
        for matrix in m_table:
            dim1=dimensions(matrix)
            
            start=dim1[0]

            for i in range(start,dim1[1]-3): #alla fine c'e' un campo LEGENDA che occupa circa 3 spazi
                h=True
                if toadd:
                    for element in toadd:
                        element_in_matrix=str(matrix[i][0])
                        if (all_replacing(element_in_matrix)==all_replacing(element)): #se l'elemento della matrice che sto controllato esiste gia' , non deve fare nulla
                            h=False
                if h:
                    element_in_matrix=str(matrix[i][0])
                    ##perche' cosi' abbiamo CATANIA ...
                    if str(matrix[i][0])=="NESIMA":
                        element_in_matrix="Catania (Metro Nesima)"
                    elif str(matrix[i][0])=="CIBALI":
                        element_in_matrix="CATANIA CIBALI"
                    toadd.append(str(matrix[i][0]))
                    b=False
                    t=False
                    for j in range(len(matrix[i])):
                        if(isTimeFormat(str(matrix[i][j])) or isTimeFormatH(str(matrix[i][j]))):
                            if tipo=="bus" and str(matrix[start-2][j])=="BUS":
                                b=True
                            elif tipo=="littorina" and str(matrix[start-2][j]).replace('.','')=="TR":
                                t=True
                    insert=element_in_matrix
                    if tipo=="bus" and b:
                        different_loc.append(insert)
                    elif tipo=="littorina" and t:
                        different_loc.append(insert)
    
    different_loc.sort()
    with open(path+tipo+"/locations.txt","w") as f:
        for element in different_loc:
            f.write(element+"\n")